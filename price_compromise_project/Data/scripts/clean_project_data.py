#!/usr/bin/env python3
"""Clean raw data into analysis-ready intermediate tables.

Outputs are written to Data/cleaned and are deliberately modular:
- BEA 2019 IO use matrix and input coefficients
- BEA annual industry price indexes
- BEA-NAICS concordance
- BLS CES and QCEW wage panels
- RTP tariff exposure file converted from Stata
- USTR Section 301 link inventory
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


YEARS = list(range(2016, 2026))


def clean_name(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def as_number(value: object) -> float:
    if value in (None, "", "..."):
        return 0.0
    try:
        return float(value)
    except Exception:
        return np.nan


def ensure_dirs(project_root: Path) -> Path:
    cleaned = project_root / "Data" / "cleaned"
    cleaned.mkdir(parents=True, exist_ok=True)
    return cleaned


def clean_bea_use_summary(project_root: Path, cleaned: Path) -> dict[str, int]:
    src = project_root / "Data" / "bea" / "raw" / "unzipped" / "SUPPLY-USE" / "Use_Summary.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["2019"]

    industry_cols = []
    for col in range(3, ws.max_column + 1):
        code = ws.cell(6, col).value
        name = ws.cell(7, col).value
        if code == "T001":
            break
        if code and name:
            industry_cols.append((col, str(code), str(name)))

    commodity_rows = []
    for row in range(8, ws.max_row + 1):
        code = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        if code == "T005":
            break
        if code and name:
            commodity_rows.append((row, str(code), str(name)))

    totals = {}
    output = {}
    va = {}
    for col, industry_code, _ in industry_cols:
        totals[industry_code] = as_number(ws.cell(81, col).value)
        va[industry_code] = as_number(ws.cell(86, col).value)
        output[industry_code] = as_number(ws.cell(87, col).value)

    long_rows = []
    for row, commodity_code, commodity_name in commodity_rows:
        for col, industry_code, industry_name in industry_cols:
            value = as_number(ws.cell(row, col).value)
            total_intermediate = totals[industry_code]
            input_share = value / total_intermediate if total_intermediate else np.nan
            long_rows.append(
                {
                    "year": 2019,
                    "commodity_code": commodity_code,
                    "commodity_name": commodity_name,
                    "industry_code": industry_code,
                    "industry_name": industry_name,
                    "value_musd": value,
                    "total_intermediate_musd": total_intermediate,
                    "input_share": input_share,
                    "link_5pct": int(input_share > 0.05) if pd.notna(input_share) else 0,
                }
            )

    long_df = pd.DataFrame(long_rows)
    long_df.to_csv(cleaned / "bea_use_summary_2019_long.csv", index=False)

    coeff = long_df[
        [
            "year",
            "industry_code",
            "industry_name",
            "commodity_code",
            "commodity_name",
            "value_musd",
            "total_intermediate_musd",
            "input_share",
            "link_5pct",
        ]
    ].copy()
    coeff.to_csv(cleaned / "bea_input_coefficients_2019.csv", index=False)

    industry_df = pd.DataFrame(
        [
            {
                "year": 2019,
                "industry_code": code,
                "industry_name": name,
                "total_intermediate_musd": totals[code],
                "value_added_basic_musd": va[code],
                "total_industry_output_basic_musd": output[code],
            }
            for _, code, name in industry_cols
        ]
    )
    industry_df.to_csv(cleaned / "bea_industry_output_2019.csv", index=False)
    return {
        "bea_use_summary_2019_long.csv": len(long_df),
        "bea_input_coefficients_2019.csv": len(coeff),
        "bea_industry_output_2019.csv": len(industry_df),
    }


def clean_bea_price_sheet(src: Path, sheet: str, output_name: str, cleaned: Path) -> int:
    raw = pd.read_excel(src, sheet_name=sheet, header=None, engine="openpyxl")
    header_row = raw.index[raw.iloc[:, 0].astype(str).str.strip().eq("Line")][0]
    year_cols = {}
    for col in raw.columns:
        val = raw.iat[header_row, col]
        try:
            year = int(val)
        except Exception:
            continue
        if year in YEARS:
            year_cols[col] = year

    rows = []
    for r in range(header_row + 1, len(raw)):
        line = raw.iat[r, 0]
        name = raw.iat[r, 1]
        if pd.isna(line) or pd.isna(name):
            continue
        for col, year in year_cols.items():
            value = raw.iat[r, col]
            value = pd.to_numeric(value, errors="coerce")
            if pd.isna(value):
                continue
            rows.append(
                {
                    "line": str(line).strip(),
                    "industry_name": str(name).strip(),
                    "industry_name_clean": clean_name(name),
                    "year": year,
                    "price_index_2017_100": float(value),
                }
            )
    df = pd.DataFrame(rows)
    df.to_csv(cleaned / output_name, index=False)
    return len(df)


def clean_bea_prices(project_root: Path, cleaned: Path) -> dict[str, int]:
    gross = project_root / "Data" / "bea" / "raw" / "GrossOutput.xlsx"
    interm = project_root / "Data" / "bea" / "raw" / "IntermediateInputs.xlsx"
    return {
        "bea_gross_output_price_index_annual_2016_2025.csv": clean_bea_price_sheet(
            gross, "TGO104-A", "bea_gross_output_price_index_annual_2016_2025.csv", cleaned
        ),
        "bea_intermediate_inputs_price_index_annual_2016_2025.csv": clean_bea_price_sheet(
            interm, "TII104-A", "bea_intermediate_inputs_price_index_annual_2016_2025.csv", cleaned
        ),
    }


def clean_bea_concordance(project_root: Path, cleaned: Path) -> int:
    src = project_root / "Data" / "bea" / "raw" / "BEA-Industry-and-Commodity-Codes-and-NAICS-Concordance.xlsx"
    raw = pd.read_excel(src, sheet_name="NAICS Codes", header=4, engine="openpyxl")
    raw = raw.rename(
        columns={
            "Sector": "sector_code",
            "Description": "sector_name",
            "Summary": "summary_code",
            "Description.1": "summary_name",
            "U. Summary": "underlying_summary_code",
            "Description.2": "underlying_summary_name",
            "Detail": "detail_code",
            "Description.3": "detail_name",
            "GO Detail": "go_detail_code",
            "Description.4": "go_detail_name",
            "Notes": "notes",
            "Related 2017 NAICS Codes": "naics_2017",
        }
    )
    cols = [
        "sector_code",
        "sector_name",
        "summary_code",
        "summary_name",
        "underlying_summary_code",
        "underlying_summary_name",
        "detail_code",
        "detail_name",
        "go_detail_code",
        "go_detail_name",
        "notes",
        "naics_2017",
    ]
    df = raw[[c for c in cols if c in raw.columns]].dropna(how="all")
    for c in df.columns:
        if "code" in c or c == "naics_2017":
            df[c] = df[c].astype("string").str.replace(r"\.0$", "", regex=True).str.strip()
    df.to_csv(cleaned / "bea_naics_concordance_clean.csv", index=False)
    return len(df)


def clean_bls(project_root: Path, cleaned: Path) -> dict[str, int]:
    ces_src = project_root / "Data" / "bls" / "ces" / "processed" / "ces_wages_monthly_2016_2025.csv"
    ces = pd.read_csv(ces_src, dtype={"industry_code": str, "month": str})
    ces["value"] = pd.to_numeric(ces["value"], errors="coerce")
    ces["year"] = pd.to_numeric(ces["year"], errors="coerce").astype("Int64")
    ces_monthly = ces.sort_values(["series_id", "year", "month"])
    ces_monthly.to_csv(cleaned / "bls_ces_wages_monthly_clean.csv", index=False)
    ces_annual = (
        ces_monthly.groupby(["series_id", "industry_code", "industry_name", "data_type", "data_type_name", "year"], dropna=False)
        .agg(avg_hourly_earnings=("value", "mean"), n_months=("value", "count"))
        .reset_index()
    )
    ces_annual.to_csv(cleaned / "bls_ces_wages_annual_2016_2025.csv", index=False)

    qcew_src = project_root / "Data" / "bls" / "qcew" / "processed" / "qcew_us_annual_private_industries_2016_2025.csv"
    qcew = pd.read_csv(qcew_src, dtype={"industry_code": str, "year": str})
    numeric_cols = [
        "annual_avg_estabs",
        "annual_avg_emplvl",
        "total_annual_wages",
        "annual_avg_wkly_wage",
        "avg_annual_pay",
    ]
    for col in numeric_cols:
        qcew[col] = pd.to_numeric(qcew[col], errors="coerce")
    qcew["year"] = pd.to_numeric(qcew["year"], errors="coerce").astype("Int64")
    qcew["naics_code"] = qcew["industry_code"].astype(str)
    qcew = qcew[
        [
            "area_fips",
            "own_code",
            "naics_code",
            "agglvl_code",
            "year",
            "annual_avg_estabs",
            "annual_avg_emplvl",
            "total_annual_wages",
            "annual_avg_wkly_wage",
            "avg_annual_pay",
        ]
    ].sort_values(["naics_code", "year"])
    qcew.to_csv(cleaned / "bls_qcew_private_industry_wages_annual_2016_2025.csv", index=False)
    return {
        "bls_ces_wages_monthly_clean.csv": len(ces_monthly),
        "bls_ces_wages_annual_2016_2025.csv": len(ces_annual),
        "bls_qcew_private_industry_wages_annual_2016_2025.csv": len(qcew),
    }


def clean_tariffs(project_root: Path, cleaned: Path) -> dict[str, int]:
    out: dict[str, int] = {}
    rtp_src = project_root / "Data" / "dataverse_files" / "rtp" / "data" / "analysis" / "tariffs_naics.dta"
    if rtp_src.exists():
        df = pd.read_stata(rtp_src)
        for c in df.columns:
            if c.startswith(("R", "S", "E", "T")):
                df[c] = pd.to_numeric(df[c], errors="coerce")
        df["naics_str"] = df["naics_str"].astype(str).str.strip()
        df.to_csv(cleaned / "rtp_tariffs_naics_clean.csv", index=False)
        out["rtp_tariffs_naics_clean.csv"] = len(df)

    links_src = project_root / "Data" / "tariffs" / "ustr_301" / "raw" / "ustr_301_links.csv"
    if links_src.exists():
        links = pd.read_csv(links_src)
        links["url_lower"] = links["url"].astype(str).str.lower()
        links["is_machine_readable"] = links["url_lower"].str.endswith((".csv", ".xlsx", ".xls", ".zip")).astype(int)
        links["is_pdf"] = links["url_lower"].str.endswith(".pdf").astype(int)
        links.to_csv(cleaned / "ustr_301_links_clean.csv", index=False)
        out["ustr_301_links_clean.csv"] = len(links)
    return out


def write_readme(cleaned: Path, counts: dict[str, int]) -> None:
    lines = [
        "# Cleaned data",
        "",
        f"Generated at {datetime.now(timezone.utc).isoformat()}",
        "",
        "Notes:",
        "- BEA IO cells marked `...` are coded as 0 in the cleaned use matrix.",
        "- `bea_input_coefficients_2019.csv` uses commodity-by-industry use values divided by each industry's total intermediate inputs.",
        "- `link_5pct` equals 1 when the input share is greater than 0.05.",
        "- QCEW wage data use national private ownership (`own_code=5`).",
        "- RTP tariff data are pre-existing Dataverse/Stata exposure variables, not a full 2018-2025 HTS tariff panel.",
        "",
        "Files:",
    ]
    for name, n in sorted(counts.items()):
        lines.append(f"- `{name}`: {n} rows")
    (cleaned / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    project_root = args.project_root.expanduser().resolve()
    cleaned = ensure_dirs(project_root)

    counts: dict[str, int] = {}
    counts.update(clean_bea_use_summary(project_root, cleaned))
    counts.update(clean_bea_prices(project_root, cleaned))
    counts["bea_naics_concordance_clean.csv"] = clean_bea_concordance(project_root, cleaned)
    counts.update(clean_bls(project_root, cleaned))
    counts.update(clean_tariffs(project_root, cleaned))
    write_readme(cleaned, counts)

    manifest = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "project_root": str(project_root),
        "counts": counts,
    }
    (cleaned / "clean_manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(manifest, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
